"""openpyxl Excel/CSV gateway (implements domain.ports.ExcelGateway).

Parses training files into ParsedTerm and builds template/export workbooks in
the same column shape (canon / th / aliases)."""
from __future__ import annotations

import csv
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from ..domain.entities import ParsedTerm

_CANON_KEYS = {"canon", "term", "head", "keyword", "คำหลัก", "หัวข้อ"}
_TH_KEYS = {"th", "thai", "desc", "description", "คำอธิบาย", "ไทย", "ความหมาย"}
_ALIAS_KEYS = {"alias", "aliases", "variant", "variants", "synonym", "คำพ้อง", "คำที่เขียนต่าง"}

_HEADER = ["canon", "th", "aliases"]
_HEADER_FILL = PatternFill("solid", fgColor="4361EE")  # PiKaOs indigo
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALIAS_HINT = 'คั่นหลายคำด้วย "|" หรือ ","'


def _split_aliases(cell: str) -> list[str]:
    if not cell:
        return []
    out = []
    for chunk in str(cell).replace("|", ",").replace(";", ",").split(","):
        c = chunk.strip()
        if c:
            out.append(c)
    return out


def _rows_from_xlsx(data: bytes) -> list[list[str]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [["" if v is None else str(v).strip() for v in r] for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _rows_from_csv(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    return [[c.strip() for c in row] for row in csv.reader(io.StringIO(text))]


class OpenpyxlExcel:
    def parse(self, filename: str, data: bytes) -> tuple[list[ParsedTerm], int]:
        name = (filename or "").lower()
        rows = _rows_from_csv(data) if name.endswith(".csv") else _rows_from_xlsx(data)
        rows = [r for r in rows if any(c for c in r)]
        if not rows:
            return [], 0

        header = [h.lower().strip() for h in rows[0]]
        ci_canon = ci_th = ci_alias = None
        for i, h in enumerate(header):
            if ci_canon is None and h in _CANON_KEYS:
                ci_canon = i
            elif ci_th is None and h in _TH_KEYS:
                ci_th = i
            elif ci_alias is None and h in _ALIAS_KEYS:
                ci_alias = i

        if ci_canon is not None:
            body = rows[1:]
        else:  # no recognizable header -> positional fallback, all rows are data
            ci_canon, ci_th, ci_alias = 0, 1, 2
            body = rows

        def cell(row: list[str], idx: int | None) -> str:
            return row[idx] if idx is not None and idx < len(row) else ""

        terms: list[ParsedTerm] = []
        for row in body:
            canon = cell(row, ci_canon)
            if not canon:
                continue
            terms.append(ParsedTerm(canon=canon, th=cell(row, ci_th) or canon, aliases=_split_aliases(cell(row, ci_alias))))
        return terms, len(body)

    # ---- workbook generation ----
    def _build(self, rows: list[ParsedTerm], sheet_title: str, with_hint: bool) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31] or "vocab"
        for col, name in enumerate(_HEADER, start=1):
            c = ws.cell(row=1, column=col, value=name)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 46
        ws.freeze_panes = "A2"
        if with_hint:
            ws.cell(row=1, column=4, value=_ALIAS_HINT).font = Font(italic=True, color="888888")
        for r, t in enumerate(rows, start=2):
            ws.cell(row=r, column=1, value=t.canon)
            ws.cell(row=r, column=2, value=t.th)
            ws.cell(row=r, column=3, value=" | ".join(t.aliases))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def build_template(self) -> bytes:
        sample = [
            ParsedTerm("Share Price", "ราคาหลักทรัพย์", ["ราคาหุ้น", "stock price"]),
            ParsedTerm("Financial Statements", "งบการเงิน", ["งบดุล", "ผลประกอบการ", "financials"]),
        ]
        return self._build(sample, "template", with_hint=True)

    def build_export(self, sheet_title: str, terms: list[ParsedTerm]) -> bytes:
        return self._build(terms, sheet_title, with_hint=False)
