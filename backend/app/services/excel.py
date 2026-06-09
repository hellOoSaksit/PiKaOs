"""Parse an uploaded Excel/CSV training file into vocabulary terms.

Expected columns (case-insensitive, Thai or English headers accepted):
  - canon / term / head / คำหลัก / หัวข้อ  -> canonical head term
  - th / thai / desc / คำอธิบาย / ไทย       -> Thai description
  - alias / aliases / variant / คำพ้อง       -> '|' or ',' separated aliases

Files without recognizable headers fall back to: col0 = canon, col1 = th,
col2 = aliases. Returns (terms, row_count).
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from openpyxl import load_workbook

_CANON_KEYS = {"canon", "term", "head", "keyword", "คำหลัก", "หัวข้อ"}
_TH_KEYS = {"th", "thai", "desc", "description", "คำอธิบาย", "ไทย", "ความหมาย"}
_ALIAS_KEYS = {"alias", "aliases", "variant", "variants", "synonym", "คำพ้อง", "คำที่เขียนต่าง"}


@dataclass
class ParsedTerm:
    canon: str
    th: str = ""
    aliases: list[str] = field(default_factory=list)


def _split_aliases(cell: str) -> list[str]:
    if not cell:
        return []
    parts = []
    for chunk in str(cell).replace("|", ",").replace(";", ",").split(","):
        c = chunk.strip()
        if c:
            parts.append(c)
    return parts


def _rows_from_xlsx(data: bytes) -> list[list[str]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v).strip() for v in r])
    wb.close()
    return rows


def _rows_from_csv(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    return [[c.strip() for c in row] for row in csv.reader(io.StringIO(text))]


def parse(filename: str, data: bytes) -> tuple[list[ParsedTerm], int]:
    name = (filename or "").lower()
    rows = _rows_from_csv(data) if name.endswith(".csv") else _rows_from_xlsx(data)
    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        return [], 0

    header = [h.lower().strip() for h in rows[0]]
    ci_canon = ci_th = ci_alias = None
    for i, h in enumerate(header):
        if ci_canon is None and h in _CANON_KEYS:
            ci_canon = i
        elif ci_th is None and h in _TH_KEYS:
            ci_th = i
        elif ci_alias is None and h in _ALIAS_KEYS:
            ci_alias = i

    if ci_canon is not None:
        body = rows[1:]
    else:  # no recognizable header -> positional fallback, treat all rows as data
        ci_canon, ci_th, ci_alias = 0, 1, 2
        body = rows

    def cell(row: list[str], idx: int | None) -> str:
        return row[idx] if idx is not None and idx < len(row) else ""

    terms: list[ParsedTerm] = []
    for row in body:
        canon = cell(row, ci_canon)
        if not canon:
            continue
        terms.append(
            ParsedTerm(
                canon=canon,
                th=cell(row, ci_th) or canon,
                aliases=_split_aliases(cell(row, ci_alias)),
            )
        )
    return terms, len(body)
